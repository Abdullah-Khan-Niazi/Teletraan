"""End-to-end tests for the order logging pipeline.

Tests the WhatsApp group logging, Excel file append, PDF generation,
and the combined logging entry point.  Uses in-memory mocks for all
database and WhatsApp operations.
"""

from __future__ import annotations

import datetime
import os
import tempfile
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import UUID, uuid4

import pytest

from app.core.constants import (
    Language,
    MedicineForm,
    OrderSource,
    OrderStatus,
    PaymentMethod,
    PaymentStatus,
)
from app.db.models.catalog import CatalogItem
from app.db.models.order import Order, OrderItem, OrderUpdate
from app.orders.logging_service import (
    _mask_phone,
    _paisas_to_pkr,
    format_whatsapp_group_message,
    log_confirmed_order,
    log_order_to_excel,
    log_order_to_whatsapp_group,
)
from app.reporting.excel_generator import (
    _format_items_column,
    _format_quantities_column,
    _get_excel_path,
    append_order_row,
    generate_daily_summary_row,
    get_monthly_report_path,
)
from app.reporting.pdf_generator import (
    generate_catalog_pdf,
    generate_order_receipt,
)


# ═══════════════════════════════════════════════════════════════════
# TEST FIXTURES
# ═══════════════════════════════════════════════════════════════════


_DIST_ID = uuid4()
_CUST_ID = uuid4()
_ORDER_ID = uuid4()
_NOW = datetime.datetime.now(datetime.timezone.utc)


def _make_order(
    *,
    status: str = OrderStatus.CONFIRMED,
    total_paisas: int = 198500,
    subtotal_paisas: int = 198500,
    discount_paisas: int = 0,
    delivery_charges_paisas: int = 0,
    payment_method: str | None = PaymentMethod.CASH,
    whatsapp_logged_at: datetime.datetime | None = None,
    excel_logged_at: datetime.datetime | None = None,
) -> Order:
    """Build a test Order."""
    return Order(
        id=_ORDER_ID,
        order_number="2847",
        distributor_id=_DIST_ID,
        customer_id=_CUST_ID,
        status=status,
        subtotal_paisas=subtotal_paisas,
        discount_paisas=discount_paisas,
        delivery_charges_paisas=delivery_charges_paisas,
        total_paisas=total_paisas,
        payment_status=PaymentStatus.UNPAID,
        payment_method=payment_method,
        delivery_address="Main Bazar, Lahore",
        source=OrderSource.WHATSAPP,
        whatsapp_logged_at=whatsapp_logged_at,
        excel_logged_at=excel_logged_at,
        created_at=_NOW,
        updated_at=_NOW,
    )


def _make_order_items() -> list[OrderItem]:
    """Build test OrderItems."""
    return [
        OrderItem(
            id=uuid4(),
            order_id=_ORDER_ID,
            distributor_id=_DIST_ID,
            medicine_name="Paracetamol 500mg",
            unit="strip",
            quantity_ordered=5,
            price_per_unit_paisas=3500,
            line_total_paisas=17500,
            discount_paisas=0,
            created_at=_NOW,
        ),
        OrderItem(
            id=uuid4(),
            order_id=_ORDER_ID,
            distributor_id=_DIST_ID,
            medicine_name="Augmentin 625mg",
            unit="box",
            quantity_ordered=2,
            price_per_unit_paisas=48000,
            line_total_paisas=96000,
            discount_paisas=5000,
            bonus_units_given=1,
            created_at=_NOW,
        ),
        OrderItem(
            id=uuid4(),
            order_id=_ORDER_ID,
            distributor_id=_DIST_ID,
            medicine_name="ORS Sachet",
            unit="carton",
            quantity_ordered=1,
            price_per_unit_paisas=85000,
            line_total_paisas=85000,
            discount_paisas=0,
            created_at=_NOW,
        ),
    ]


def _make_catalog_items() -> list[CatalogItem]:
    """Build test CatalogItems for PDF generation."""
    base = {
        "distributor_id": _DIST_ID,
        "price_per_unit_paisas": 3500,
        "created_at": _NOW,
        "updated_at": _NOW,
    }
    return [
        CatalogItem(
            id=uuid4(),
            medicine_name="Paracetamol 500mg",
            generic_name="Paracetamol",
            category="Analgesics",
            form=MedicineForm.TABLET,
            strength="500mg",
            unit="strip",
            is_in_stock=True,
            stock_quantity=200,
            **base,
        ),
        CatalogItem(
            id=uuid4(),
            medicine_name="Augmentin 625mg",
            generic_name="Amoxicillin + Clavulanate",
            category="Antibiotics",
            form=MedicineForm.TABLET,
            strength="625mg",
            unit="box",
            price_per_unit_paisas=48000,
            is_in_stock=True,
            stock_quantity=50,
            distributor_id=_DIST_ID,
            created_at=_NOW,
            updated_at=_NOW,
        ),
        CatalogItem(
            id=uuid4(),
            medicine_name="ORS Sachet",
            category="OTC",
            unit="carton",
            price_per_unit_paisas=85000,
            is_in_stock=False,
            stock_quantity=0,
            distributor_id=_DIST_ID,
            created_at=_NOW,
            updated_at=_NOW,
        ),
    ]


def _make_distributor_mock(
    *,
    group_id: str | None = "group_jid_123",
    phone_number_id: str = "1234567890",
) -> MagicMock:
    """Build a mock distributor."""
    dist = MagicMock()
    dist.id = _DIST_ID
    dist.business_name = "Khan Distributors"
    dist.whatsapp_phone_number_id = phone_number_id
    dist.whatsapp_group_id = group_id
    return dist


def _make_customer_mock() -> MagicMock:
    """Build a mock customer."""
    cust = MagicMock()
    cust.name = "Rizwan Ahmed"
    cust.shop_name = "Rizwan Medical Store"
    cust.whatsapp_number = "+923121234567"
    cust.address = "Main Bazar, Dukaan 14, Gujranwala"
    cust.city = "Gujranwala"
    return cust


# ═══════════════════════════════════════════════════════════════════
# FORMATTING TESTS
# ═══════════════════════════════════════════════════════════════════


class TestFormatHelpers:
    """Test format helper functions."""

    def test_paisas_to_pkr_whole(self):
        assert _paisas_to_pkr(198500) == "PKR 1,985"

    def test_paisas_to_pkr_fractional(self):
        assert _paisas_to_pkr(1950) == "PKR 19.50"

    def test_paisas_to_pkr_zero(self):
        assert _paisas_to_pkr(0) == "PKR 0"

    def test_mask_phone(self):
        assert _mask_phone("+923121234567") == "****4567"

    def test_mask_phone_short(self):
        assert _mask_phone("12") == "****"


class TestWhatsAppGroupFormat:
    """Test the WhatsApp group message formatting."""

    def test_format_basic_order(self):
        order = _make_order()
        items = _make_order_items()
        msg = format_whatsapp_group_message(
            order=order,
            items=items,
            customer_name="Rizwan Ahmed",
            shop_name="Rizwan Medical Store",
            customer_phone="+923121234567",
            address="Main Bazar, Gujranwala",
            city="Gujranwala",
        )

        assert "NAYA ORDER" in msg
        assert "#2847" in msg
        assert "Rizwan Ahmed" in msg
        assert "Rizwan Medical Store" in msg
        assert "Paracetamol 500mg" in msg
        assert "Augmentin 625mg" in msg
        assert "ORS Sachet" in msg
        assert "PKR 1,985" in msg
        assert "CONFIRMED" in msg
        # Phone should be masked
        assert "****4567" in msg
        assert "+923121234567" not in msg

    def test_format_with_discount(self):
        order = _make_order(discount_paisas=10000)
        items = _make_order_items()
        msg = format_whatsapp_group_message(
            order=order,
            items=items,
            customer_name="Test",
            shop_name="Test Shop",
            customer_phone="+921111111111",
            address="Addr",
            city="City",
        )
        assert "Discount" in msg

    def test_format_with_delivery_charges(self):
        order = _make_order(delivery_charges_paisas=20000)
        items = _make_order_items()
        msg = format_whatsapp_group_message(
            order=order,
            items=items,
            customer_name="Test",
            shop_name="Test Shop",
            customer_phone="+921111111111",
            address="Addr",
            city="City",
        )
        assert "Delivery" in msg

    def test_item_bonus_displayed(self):
        order = _make_order()
        items = _make_order_items()
        msg = format_whatsapp_group_message(
            order=order,
            items=items,
            customer_name="Test",
            shop_name="Test Shop",
            customer_phone="+921111111111",
            address="Addr",
            city="City",
        )
        # Augmentin has bonus_units_given=1
        assert "Bonus" in msg


# ═══════════════════════════════════════════════════════════════════
# EXCEL GENERATOR TESTS
# ═══════════════════════════════════════════════════════════════════


class TestExcelGenerator:
    """Test Excel file generation and row append."""

    def test_format_items_column(self):
        items = _make_order_items()
        result = _format_items_column(items)
        assert "Paracetamol 500mg" in result
        assert "Augmentin 625mg" in result

    def test_format_quantities_column(self):
        items = _make_order_items()
        result = _format_quantities_column(items)
        assert "5" in result
        assert "2" in result
        assert "1" in result

    @pytest.mark.asyncio
    async def test_append_order_row_creates_file(self):
        order = _make_order()
        items = _make_order_items()

        success = await append_order_row(
            order=order,
            items=items,
            customer_name="Test Customer",
            shop_name="Test Shop",
            city="Lahore",
        )

        assert success is True

        # Verify file exists
        filepath = _get_excel_path(
            str(order.distributor_id),
            _NOW.year,
            _NOW.month,
        )
        assert filepath.exists()

        # Clean up
        filepath.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_append_order_row_adds_data(self):
        order = _make_order()
        items = _make_order_items()

        await append_order_row(
            order=order,
            items=items,
            customer_name="Test Customer",
            shop_name="Test Shop",
            city="Lahore",
        )

        filepath = _get_excel_path(
            str(order.distributor_id),
            _NOW.year,
            _NOW.month,
        )

        # Verify content using openpyxl
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        # Row 1 = headers, Row 2 = data
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 1

        data_row = rows[0]
        assert data_row[0] == "2847"  # Order number
        assert data_row[3] == "Test Customer"
        assert data_row[4] == "Test Shop"

        wb.close()
        filepath.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_append_multiple_rows(self):
        order = _make_order()
        items = _make_order_items()

        await append_order_row(order, items, "Cust 1", "Shop 1", "City 1")
        await append_order_row(order, items, "Cust 2", "Shop 2", "City 2")

        filepath = _get_excel_path(
            str(order.distributor_id),
            _NOW.year,
            _NOW.month,
        )

        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        wb.close()
        filepath.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_get_monthly_report_path_exists(self):
        order = _make_order()
        items = _make_order_items()

        await append_order_row(order, items, "Test", "Shop", "City")

        path = await get_monthly_report_path(
            str(order.distributor_id),
            _NOW.year,
            _NOW.month,
        )
        assert path is not None
        assert path.exists()

        path.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_get_monthly_report_path_not_exists(self):
        path = await get_monthly_report_path(
            "nonexistent_id", 2020, 1,
        )
        assert path is None


# ═══════════════════════════════════════════════════════════════════
# PDF GENERATOR TESTS
# ═══════════════════════════════════════════════════════════════════


class TestPDFGenerator:
    """Test PDF generation for catalog and receipt."""

    @pytest.mark.asyncio
    async def test_generate_catalog_pdf(self):
        items = _make_catalog_items()
        path = await generate_catalog_pdf(
            items=items,
            distributor_id=str(_DIST_ID),
            distributor_name="Khan Distributors",
        )

        assert path is not None
        assert path.exists()
        assert path.suffix == ".pdf"

        # PDF should be non-empty
        assert path.stat().st_size > 1000

        path.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_generate_catalog_pdf_empty_items(self):
        path = await generate_catalog_pdf(
            items=[],
            distributor_id=str(_DIST_ID),
            distributor_name="Empty Catalog Dist",
        )

        assert path is not None
        assert path.exists()
        path.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_generate_order_receipt(self):
        order = _make_order()
        items = _make_order_items()

        path = await generate_order_receipt(
            order=order,
            items=items,
            customer_name="Rizwan Ahmed",
            shop_name="Rizwan Medical Store",
            address="Main Bazar, Gujranwala",
            city="Gujranwala",
            customer_phone="+923121234567",
            distributor_name="Khan Distributors",
        )

        assert path is not None
        assert path.exists()
        assert path.suffix == ".pdf"
        assert path.stat().st_size > 1000

        path.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_receipt_has_order_number_in_filename(self):
        order = _make_order()
        items = _make_order_items()

        path = await generate_order_receipt(
            order=order,
            items=items,
            customer_name="Test",
            shop_name="Test Shop",
            address="Addr",
            city="City",
            customer_phone="+921111111111",
            distributor_name="Test Dist",
        )

        assert path is not None
        assert "2847" in path.name
        path.unlink(missing_ok=True)


# ═══════════════════════════════════════════════════════════════════
# WHATSAPP GROUP LOGGING TESTS
# ═══════════════════════════════════════════════════════════════════


class TestWhatsAppGroupLogging:
    """Test the WhatsApp group logging service."""

    @pytest.mark.asyncio
    async def test_log_sends_to_group(self):
        order = _make_order()
        items = _make_order_items()

        dist_repo = AsyncMock()
        dist_repo.get_by_id.return_value = _make_distributor_mock()

        cust_repo = AsyncMock()
        cust_repo.get_by_id.return_value = _make_customer_mock()

        order_repo = AsyncMock()

        notifier = AsyncMock()
        notifier.send_text.return_value = "msg_id_123"

        result = await log_order_to_whatsapp_group(
            order, items,
            order_repo=order_repo,
            customer_repo=cust_repo,
            distributor_repo=dist_repo,
            notifier=notifier,
        )

        assert result is True
        notifier.send_text.assert_called_once()
        call_kwargs = notifier.send_text.call_args
        assert "group_jid_123" in str(call_kwargs)
        order_repo.update.assert_called_once()

    @pytest.mark.asyncio
    async def test_skip_if_already_logged(self):
        order = _make_order(whatsapp_logged_at=_NOW)
        items = _make_order_items()

        result = await log_order_to_whatsapp_group(order, items)
        assert result is True

    @pytest.mark.asyncio
    async def test_skip_if_no_group_id(self):
        order = _make_order()
        items = _make_order_items()

        dist_repo = AsyncMock()
        dist_repo.get_by_id.return_value = _make_distributor_mock(group_id=None)

        result = await log_order_to_whatsapp_group(
            order, items,
            distributor_repo=dist_repo,
        )
        assert result is False

    @pytest.mark.asyncio
    async def test_handles_send_failure(self):
        order = _make_order()
        items = _make_order_items()

        dist_repo = AsyncMock()
        dist_repo.get_by_id.return_value = _make_distributor_mock()

        cust_repo = AsyncMock()
        cust_repo.get_by_id.return_value = _make_customer_mock()

        order_repo = AsyncMock()
        notifier = AsyncMock()
        notifier.send_text.return_value = None  # send failed

        result = await log_order_to_whatsapp_group(
            order, items,
            order_repo=order_repo,
            customer_repo=cust_repo,
            distributor_repo=dist_repo,
            notifier=notifier,
        )

        assert result is False
        order_repo.update.assert_not_called()


# ═══════════════════════════════════════════════════════════════════
# EXCEL LOGGING SERVICE TESTS
# ═══════════════════════════════════════════════════════════════════


class TestExcelLogging:
    """Test the Excel logging service entry point."""

    @pytest.mark.asyncio
    async def test_log_to_excel_succeeds(self):
        order = _make_order()
        items = _make_order_items()

        cust_repo = AsyncMock()
        cust_repo.get_by_id.return_value = _make_customer_mock()

        order_repo = AsyncMock()

        with patch("app.orders.logging_service.get_settings") as mock_settings:
            mock_settings.return_value.enable_excel_reports = True
            result = await log_order_to_excel(
                order, items,
                order_repo=order_repo,
                customer_repo=cust_repo,
            )

        assert result is True
        order_repo.update.assert_called_once()

        # Clean up generated file
        filepath = _get_excel_path(
            str(order.distributor_id), _NOW.year, _NOW.month,
        )
        filepath.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_skip_if_disabled(self):
        order = _make_order()
        items = _make_order_items()

        with patch("app.orders.logging_service.get_settings") as mock_settings:
            mock_settings.return_value.enable_excel_reports = False
            result = await log_order_to_excel(order, items)

        assert result is False

    @pytest.mark.asyncio
    async def test_skip_if_already_logged(self):
        order = _make_order(excel_logged_at=_NOW)
        items = _make_order_items()

        with patch("app.orders.logging_service.get_settings") as mock_settings:
            mock_settings.return_value.enable_excel_reports = True
            result = await log_order_to_excel(order, items)

        assert result is True


# ═══════════════════════════════════════════════════════════════════
# COMBINED LOGGING TESTS
# ═══════════════════════════════════════════════════════════════════


class TestCombinedLogging:
    """Test the combined log_confirmed_order entry point."""

    @pytest.mark.asyncio
    async def test_both_destinations(self):
        order = _make_order()
        items = _make_order_items()

        order_repo = AsyncMock()
        order_repo.get_order_with_items.return_value = (order, items)

        dist_repo = AsyncMock()
        dist_repo.get_by_id.return_value = _make_distributor_mock()

        cust_repo = AsyncMock()
        cust_repo.get_by_id.return_value = _make_customer_mock()

        notifier = AsyncMock()
        notifier.send_text.return_value = "msg_id_456"

        with patch("app.orders.logging_service.get_settings") as mock_settings:
            mock_settings.return_value.enable_excel_reports = True
            results = await log_confirmed_order(
                str(_ORDER_ID),
                str(_DIST_ID),
                order_repo=order_repo,
                customer_repo=cust_repo,
                distributor_repo=dist_repo,
                notifier=notifier,
            )

        assert results["whatsapp"] is True
        assert results["excel"] is True

        # Clean up
        filepath = _get_excel_path(
            str(_DIST_ID), _NOW.year, _NOW.month,
        )
        filepath.unlink(missing_ok=True)

    @pytest.mark.asyncio
    async def test_handles_fetch_failure(self):
        order_repo = AsyncMock()
        order_repo.get_order_with_items.side_effect = Exception("DB error")

        results = await log_confirmed_order(
            str(_ORDER_ID),
            str(_DIST_ID),
            order_repo=order_repo,
        )

        assert results["whatsapp"] is False
        assert results["excel"] is False

    @pytest.mark.asyncio
    async def test_whatsapp_fails_excel_succeeds(self):
        order = _make_order()
        items = _make_order_items()

        order_repo = AsyncMock()
        order_repo.get_order_with_items.return_value = (order, items)

        dist_repo = AsyncMock()
        dist_repo.get_by_id.return_value = _make_distributor_mock(group_id=None)

        cust_repo = AsyncMock()
        cust_repo.get_by_id.return_value = _make_customer_mock()

        notifier = AsyncMock()

        with patch("app.orders.logging_service.get_settings") as mock_settings:
            mock_settings.return_value.enable_excel_reports = True
            results = await log_confirmed_order(
                str(_ORDER_ID),
                str(_DIST_ID),
                order_repo=order_repo,
                customer_repo=cust_repo,
                distributor_repo=dist_repo,
                notifier=notifier,
            )

        # WhatsApp fails (no group_id), Excel succeeds
        assert results["whatsapp"] is False
        assert results["excel"] is True

        # Clean up
        filepath = _get_excel_path(
            str(_DIST_ID), _NOW.year, _NOW.month,
        )
        filepath.unlink(missing_ok=True)
