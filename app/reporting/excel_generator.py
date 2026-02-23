"""Monthly Excel order log generator using openpyxl.

Maintains one ``.xlsx`` file per distributor per month.  Each confirmed
order is appended as a new row containing: order number, date/time,
customer, shop, items, quantities, unit prices, discounts, total,
payment method, delivery zone, city, and status.

File naming: ``orders_{distributor_id}_{YYYY_MM}.xlsx``

Files are stored in a temporary directory and can be uploaded to
Supabase Storage or emailed to the distributor by the scheduler.
"""

from __future__ import annotations

import asyncio
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from loguru import logger

from app.db.models.order import Order, OrderItem


# ═══════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════


_REPORTS_DIR = Path(tempfile.gettempdir()) / "teletraan_reports"

_HEADER_COLUMNS: list[str] = [
    "Order #",
    "Date",
    "Time",
    "Customer Name",
    "Shop Name",
    "City",
    "Items",
    "Quantities",
    "Unit Prices (PKR)",
    "Discounts (PKR)",
    "Subtotal (PKR)",
    "Delivery (PKR)",
    "Total (PKR)",
    "Payment Method",
    "Status",
]

_HEADER_WIDTHS: list[int] = [
    12,   # Order #
    12,   # Date
    8,    # Time
    20,   # Customer Name
    22,   # Shop Name
    14,   # City
    40,   # Items
    14,   # Quantities
    18,   # Unit Prices
    16,   # Discounts
    14,   # Subtotal
    14,   # Delivery
    14,   # Total
    16,   # Payment Method
    14,   # Status
]


# ═══════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════


def _paisas_to_pkr_decimal(paisas: int) -> float:
    """Convert paisas to PKR as a float for Excel numeric cells.

    Args:
        paisas: Amount in paisas.

    Returns:
        PKR value as float.
    """
    return paisas / 100.0


def _ensure_reports_dir() -> Path:
    """Create the reports directory if it does not exist.

    Returns:
        Path to the reports directory.
    """
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return _REPORTS_DIR


def _get_excel_path(distributor_id: str, year: int, month: int) -> Path:
    """Get the path of the Excel file for a distributor+month.

    Args:
        distributor_id: Distributor UUID string.
        year: Calendar year.
        month: Calendar month (1-12).

    Returns:
        Path to the Excel file.
    """
    filename = f"orders_{distributor_id}_{year:04d}_{month:02d}.xlsx"
    return _ensure_reports_dir() / filename


def _format_items_column(items: list[OrderItem]) -> str:
    """Format all item names into a comma-separated string.

    Args:
        items: Order items.

    Returns:
        String e.g. ``"Paracetamol 500mg, Augmentin 625mg"``.
    """
    return ", ".join(
        f"{item.medicine_name}{f' ({item.unit})' if item.unit else ''}"
        for item in items
    )


def _format_quantities_column(items: list[OrderItem]) -> str:
    """Format all quantities into a comma-separated string.

    Args:
        items: Order items.

    Returns:
        String e.g. ``"5, 2, 1"``.
    """
    return ", ".join(str(item.quantity_ordered) for item in items)


def _format_unit_prices_column(items: list[OrderItem]) -> str:
    """Format all unit prices into a comma-separated string.

    Args:
        items: Order items.

    Returns:
        String e.g. ``"35.00, 480.00, 850.00"``.
    """
    return ", ".join(
        f"{_paisas_to_pkr_decimal(item.price_per_unit_paisas):.2f}"
        for item in items
    )


def _format_discounts_column(items: list[OrderItem]) -> str:
    """Format per-item discounts into a comma-separated string.

    Args:
        items: Order items.

    Returns:
        String e.g. ``"0.00, 50.00, 0.00"``.
    """
    return ", ".join(
        f"{_paisas_to_pkr_decimal(item.discount_paisas):.2f}"
        for item in items
    )


# ═══════════════════════════════════════════════════════════════════
# WORKBOOK MANAGEMENT
# ═══════════════════════════════════════════════════════════════════


def _create_workbook_with_headers(filepath: Path) -> None:
    """Create a new Excel workbook with column headers and styling.

    Args:
        filepath: Where to save the new workbook.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Log"

    # Header styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_idx, (header, width) in enumerate(
        zip(_HEADER_COLUMNS, _HEADER_WIDTHS, strict=True), start=1,
    ):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # Set column width
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.debug("excel.workbook_created", path=str(filepath))


def _append_row_sync(
    filepath: Path,
    row_data: list[Any],
) -> bool:
    """Append a single row to an existing Excel workbook (sync).

    This is a blocking I/O operation — must be called via
    ``asyncio.to_thread()``.

    Args:
        filepath: Path to the Excel file.
        row_data: List of cell values to append.

    Returns:
        ``True`` on success, ``False`` on failure.
    """
    import openpyxl

    try:
        if not filepath.exists():
            _create_workbook_with_headers(filepath)

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws.append(row_data)
        wb.save(filepath)
        return True

    except Exception as exc:
        logger.error(
            "excel.append_failed",
            path=str(filepath),
            error=str(exc),
        )
        return False


# ═══════════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════════


async def append_order_row(
    order: Order,
    items: list[OrderItem],
    customer_name: str,
    shop_name: str,
    city: str,
) -> bool:
    """Append an order as a new row in the monthly Excel log.

    Creates the workbook with headers if it does not yet exist.
    Uses ``asyncio.to_thread()`` to avoid blocking the event loop
    during file I/O.

    Args:
        order: The confirmed order.
        items: Line items for the order.
        customer_name: Customer full name.
        shop_name: Shop / pharmacy name.
        city: Customer city.

    Returns:
        ``True`` if the row was appended successfully, ``False`` otherwise.
    """
    confirmed_at = order.updated_at or order.created_at
    year = confirmed_at.year
    month = confirmed_at.month

    filepath = _get_excel_path(str(order.distributor_id), year, month)

    # Status display
    status = (
        order.status.value
        if hasattr(order.status, "value")
        else str(order.status)
    ).upper()

    # Payment method
    payment = ""
    if order.payment_method:
        payment = (
            order.payment_method.value
            if hasattr(order.payment_method, "value")
            else str(order.payment_method)
        ).replace("_", " ").title()

    row_data: list[Any] = [
        order.order_number,
        confirmed_at.strftime("%d-%b-%y"),
        confirmed_at.strftime("%H:%M"),
        customer_name,
        shop_name,
        city,
        _format_items_column(items),
        _format_quantities_column(items),
        _format_unit_prices_column(items),
        _format_discounts_column(items),
        _paisas_to_pkr_decimal(order.subtotal_paisas),
        _paisas_to_pkr_decimal(order.delivery_charges_paisas),
        _paisas_to_pkr_decimal(order.total_paisas),
        payment,
        status,
    ]

    success = await asyncio.to_thread(_append_row_sync, filepath, row_data)

    if success:
        logger.info(
            "excel.order_appended",
            order_number=order.order_number,
            file=filepath.name,
        )

    return success


async def get_monthly_report_path(
    distributor_id: str,
    year: int | None = None,
    month: int | None = None,
) -> Path | None:
    """Get the file path for a distributor's monthly Excel report.

    If year/month are not specified, defaults to the current month.

    Args:
        distributor_id: Distributor UUID string.
        year: Calendar year.
        month: Calendar month (1-12).

    Returns:
        Path to the Excel file if it exists, ``None`` otherwise.
    """
    now = datetime.now(timezone.utc)
    year = year or now.year
    month = month or now.month

    filepath = _get_excel_path(distributor_id, year, month)

    if filepath.exists():
        return filepath

    logger.debug(
        "excel.report_not_found",
        distributor_id=distributor_id,
        year=year,
        month=month,
    )
    return None


async def generate_daily_summary_row(
    distributor_id: str,
) -> dict[str, Any]:
    """Compute a daily summary from the current month's Excel file.

    Reads the current month's workbook and computes totals for today.

    Args:
        distributor_id: Distributor UUID string.

    Returns:
        Dict with ``total_orders``, ``total_revenue_pkr``, ``date``.
        Empty dict if no data found.
    """
    now = datetime.now(timezone.utc)
    filepath = _get_excel_path(distributor_id, now.year, now.month)

    if not filepath.exists():
        return {}

    today_str = now.strftime("%d-%b-%y")

    def _compute_sync() -> dict[str, Any]:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        total_orders = 0
        total_revenue = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 13 and row[1] == today_str:
                total_orders += 1
                try:
                    total_revenue += float(row[12]) if row[12] else 0.0
                except (ValueError, TypeError):
                    pass

        wb.close()
        return {
            "total_orders": total_orders,
            "total_revenue_pkr": total_revenue,
            "date": today_str,
        }

    try:
        return await asyncio.to_thread(_compute_sync)
    except Exception as exc:
        logger.error(
            "excel.summary_failed",
            distributor_id=distributor_id,
            error=str(exc),
        )
        return {}
