"""Inventory sync service — Excel/CSV file sync with validation and upsert.

Reads inventory data from Google Drive URLs or Supabase Storage uploads,
validates columns and data types, upserts into the catalog table, and
writes comprehensive logs to ``inventory_sync_log``.

Sync flow:
    1. Start a sync log entry (status = STARTED).
    2. Download the file (Google Drive or Supabase Storage).
    3. Parse as Excel (.xlsx) or CSV.
    4. Validate required columns exist.
    5. For each valid row, upsert into catalog (match by SKU → medicine_name).
    6. Collect per-row errors for invalid rows.
    7. Update sync log with final counts and status.
    8. Invalidate catalog cache for the distributor.
    9. Return sync results.
"""

from __future__ import annotations

import asyncio
import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, Sequence
from uuid import UUID

import httpx
from loguru import logger

from app.core.constants import SyncSource, SyncStatus
from app.core.exceptions import DatabaseError, ValidationError
from app.db.models.audit import (
    InventorySyncLog,
    InventorySyncLogCreate,
    InventorySyncLogUpdate,
)
from app.db.models.catalog import CatalogItemCreate
from app.db.repositories import catalog_repo, sync_log_repo
from app.inventory.catalog_service import CatalogService


# ═══════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════

# Required columns in the uploaded file (case-insensitive match)
REQUIRED_COLUMNS: frozenset[str] = frozenset({
    "medicine_name",
    "price_per_unit_paisas",
    "stock_quantity",
})

# All recognized columns (used for mapping)
KNOWN_COLUMNS: frozenset[str] = frozenset({
    "medicine_name",
    "generic_name",
    "brand_name",
    "manufacturer",
    "category",
    "form",
    "strength",
    "unit",
    "units_per_pack",
    "price_per_unit_paisas",
    "mrp_paisas",
    "stock_quantity",
    "low_stock_threshold",
    "allow_order_when_out_of_stock",
    "requires_prescription",
    "is_controlled_substance",
    "search_keywords",
    "sku",
    "barcode",
    "image_url",
})

# Common column aliases → canonical name mapping
_COLUMN_ALIASES: dict[str, str] = {
    "name": "medicine_name",
    "medicine": "medicine_name",
    "med_name": "medicine_name",
    "item_name": "medicine_name",
    "generic": "generic_name",
    "brand": "brand_name",
    "mfr": "manufacturer",
    "mfg": "manufacturer",
    "cat": "category",
    "dosage_form": "form",
    "price": "price_per_unit_paisas",
    "price_paisas": "price_per_unit_paisas",
    "unit_price": "price_per_unit_paisas",
    "mrp": "mrp_paisas",
    "stock": "stock_quantity",
    "qty": "stock_quantity",
    "quantity": "stock_quantity",
    "low_stock": "low_stock_threshold",
    "threshold": "low_stock_threshold",
    "keywords": "search_keywords",
}

# Google Drive direct download URL pattern
_GDRIVE_FILE_ID_RE = re.compile(
    r"(?:/d/|id=|open\?id=)([a-zA-Z0-9_-]{25,})"
)

# HTTP timeout for file downloads
_DOWNLOAD_TIMEOUT = 60.0

# Maximum file size: 10 MB
_MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024


# ═══════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════


@dataclass(slots=True)
class RowError:
    """Describes a single row-level validation or upsert error."""

    row_number: int
    column: Optional[str] = None
    value: Optional[str] = None
    error: str = ""


@dataclass(slots=True)
class SyncResult:
    """Summary of an inventory sync operation."""

    sync_log_id: str
    distributor_id: str
    status: SyncStatus
    rows_processed: int = 0
    rows_inserted: int = 0
    rows_updated: int = 0
    rows_failed: int = 0
    errors: list[RowError] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════
# SERVICE
# ═══════════════════════════════════════════════════════════════════


class InventorySyncService:
    """Orchestrates inventory file sync for a distributor.

    Handles downloading, parsing, validating, and upserting catalog
    data from Excel (.xlsx) or CSV files.
    """

    def __init__(
        self,
        catalog_svc: CatalogService | None = None,
    ) -> None:
        self._catalog_svc = catalog_svc or CatalogService()

    # ── Public API ──────────────────────────────────────────────────

    async def sync_from_url(
        self,
        distributor_id: str,
        file_url: str,
        *,
        sync_source: SyncSource = SyncSource.GOOGLE_DRIVE,
    ) -> SyncResult:
        """Download a file from a URL and sync its contents to the catalog.

        Supports Google Drive shareable links and direct HTTP URLs to
        ``.xlsx`` or ``.csv`` files.

        Args:
            distributor_id: Tenant UUID string.
            file_url: URL to the inventory file.
            sync_source: Origin of the file (default GOOGLE_DRIVE).

        Returns:
            SyncResult with counts and any per-row errors.

        Raises:
            ValidationError: If the URL is unreachable or the file is invalid.
        """
        file_name = self._extract_filename_from_url(file_url)
        log_entry = await self._start_sync_log(
            distributor_id=distributor_id,
            sync_source=sync_source,
            file_name=file_name,
            file_url=file_url,
        )
        sync_log_id = str(log_entry.id)

        try:
            file_bytes = await self._download_file(file_url)
            return await self._process_file(
                distributor_id=distributor_id,
                sync_log_id=sync_log_id,
                file_bytes=file_bytes,
                file_name=file_name,
            )
        except Exception as exc:
            logger.error(
                "inventory.sync_from_url.failed",
                distributor_id=distributor_id,
                error=str(exc),
            )
            await self._complete_sync_log(
                sync_log_id=sync_log_id,
                status=SyncStatus.FAILED,
                rows_processed=0,
                rows_inserted=0,
                rows_updated=0,
                rows_failed=0,
                errors=[RowError(row_number=0, error=str(exc))],
            )
            return SyncResult(
                sync_log_id=sync_log_id,
                distributor_id=distributor_id,
                status=SyncStatus.FAILED,
                errors=[RowError(row_number=0, error=str(exc))],
            )

    async def sync_from_bytes(
        self,
        distributor_id: str,
        file_bytes: bytes,
        file_name: str,
        *,
        sync_source: SyncSource = SyncSource.SUPABASE_UPLOAD,
    ) -> SyncResult:
        """Sync catalog from raw file bytes (e.g. Supabase Storage upload).

        Args:
            distributor_id: Tenant UUID string.
            file_bytes: Raw file content.
            file_name: Original filename (used to determine format).
            sync_source: Origin of the file (default SUPABASE_UPLOAD).

        Returns:
            SyncResult with counts and any per-row errors.

        Raises:
            ValidationError: If the file format or content is invalid.
        """
        log_entry = await self._start_sync_log(
            distributor_id=distributor_id,
            sync_source=sync_source,
            file_name=file_name,
            file_url=None,
        )
        sync_log_id = str(log_entry.id)

        try:
            return await self._process_file(
                distributor_id=distributor_id,
                sync_log_id=sync_log_id,
                file_bytes=file_bytes,
                file_name=file_name,
            )
        except Exception as exc:
            logger.error(
                "inventory.sync_from_bytes.failed",
                distributor_id=distributor_id,
                error=str(exc),
            )
            await self._complete_sync_log(
                sync_log_id=sync_log_id,
                status=SyncStatus.FAILED,
                rows_processed=0,
                rows_inserted=0,
                rows_updated=0,
                rows_failed=0,
                errors=[RowError(row_number=0, error=str(exc))],
            )
            return SyncResult(
                sync_log_id=sync_log_id,
                distributor_id=distributor_id,
                status=SyncStatus.FAILED,
                errors=[RowError(row_number=0, error=str(exc))],
            )

    async def sync_from_manual_data(
        self,
        distributor_id: str,
        rows: list[dict[str, Any]],
    ) -> SyncResult:
        """Sync catalog from pre-parsed row dicts (e.g. API upload).

        Args:
            distributor_id: Tenant UUID string.
            rows: List of dicts with catalog fields already parsed.

        Returns:
            SyncResult with counts and any per-row errors.
        """
        log_entry = await self._start_sync_log(
            distributor_id=distributor_id,
            sync_source=SyncSource.MANUAL_API,
            file_name="api_upload",
            file_url=None,
        )
        sync_log_id = str(log_entry.id)

        try:
            return await self._upsert_rows(
                distributor_id=distributor_id,
                sync_log_id=sync_log_id,
                rows=rows,
            )
        except Exception as exc:
            logger.error(
                "inventory.sync_from_manual.failed",
                distributor_id=distributor_id,
                error=str(exc),
            )
            await self._complete_sync_log(
                sync_log_id=sync_log_id,
                status=SyncStatus.FAILED,
                rows_processed=0,
                rows_inserted=0,
                rows_updated=0,
                rows_failed=0,
                errors=[RowError(row_number=0, error=str(exc))],
            )
            return SyncResult(
                sync_log_id=sync_log_id,
                distributor_id=distributor_id,
                status=SyncStatus.FAILED,
                errors=[RowError(row_number=0, error=str(exc))],
            )

    # ── File Download ───────────────────────────────────────────────

    async def _download_file(self, url: str) -> bytes:
        """Download a file from URL with Google Drive support.

        Converts Google Drive shareable links to direct download URLs.
        Enforces a 10 MB size limit.

        Args:
            url: HTTP(S) URL to download.

        Returns:
            Raw file bytes.

        Raises:
            ValidationError: If URL is unreachable, too large, or
                returns a non-2xx status.
        """
        download_url = self._convert_gdrive_url(url)

        try:
            async with httpx.AsyncClient(
                timeout=_DOWNLOAD_TIMEOUT,
                follow_redirects=True,
            ) as client:
                response = await client.get(download_url)
                response.raise_for_status()

                if len(response.content) > _MAX_FILE_SIZE_BYTES:
                    raise ValidationError(
                        f"File exceeds maximum size of {_MAX_FILE_SIZE_BYTES // (1024 * 1024)} MB",
                        operation="download_file",
                    )

                logger.info(
                    "inventory.file_downloaded",
                    url_suffix=url[-30:],
                    size_bytes=len(response.content),
                )
                return response.content

        except httpx.HTTPStatusError as exc:
            raise ValidationError(
                f"HTTP {exc.response.status_code} downloading inventory file",
                operation="download_file",
            ) from exc
        except httpx.RequestError as exc:
            raise ValidationError(
                f"Network error downloading inventory file: {exc}",
                operation="download_file",
            ) from exc

    # ── File Parsing ────────────────────────────────────────────────

    async def _process_file(
        self,
        distributor_id: str,
        sync_log_id: str,
        file_bytes: bytes,
        file_name: str,
    ) -> SyncResult:
        """Parse a file and upsert its rows into the catalog.

        Args:
            distributor_id: Tenant UUID.
            sync_log_id: Active sync log entry ID.
            file_bytes: Raw file content.
            file_name: Original filename (determines parse format).

        Returns:
            SyncResult with full counts.

        Raises:
            ValidationError: If file format is unsupported or
                required columns are missing.
        """
        lower_name = file_name.lower() if file_name else ""

        if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            rows = await asyncio.to_thread(
                self._parse_excel, file_bytes
            )
        elif lower_name.endswith(".csv"):
            rows = self._parse_csv(file_bytes)
        else:
            raise ValidationError(
                f"Unsupported file format: {file_name}. "
                "Expected .xlsx, .xls, or .csv",
                operation="process_file",
            )

        if not rows:
            raise ValidationError(
                "File contains no data rows",
                operation="process_file",
            )

        return await self._upsert_rows(
            distributor_id=distributor_id,
            sync_log_id=sync_log_id,
            rows=rows,
        )

    def _parse_excel(self, file_bytes: bytes) -> list[dict[str, Any]]:
        """Parse Excel file bytes into a list of row dicts.

        Runs in a thread to avoid blocking the event loop (openpyxl
        is synchronous I/O).

        Args:
            file_bytes: Raw .xlsx content.

        Returns:
            List of dicts with normalized column keys.

        Raises:
            ValidationError: If openpyxl cannot read the file or
                required columns are missing.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(
                filename=io.BytesIO(file_bytes),
                read_only=True,
                data_only=True,
            )
            ws = wb.active
            if ws is None:
                raise ValidationError(
                    "Excel file has no active worksheet",
                    operation="parse_excel",
                )

            # Read header row
            raw_headers: list[str] = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = str(cell.value).strip() if cell.value is not None else ""
                raw_headers.append(val)

            headers = self._normalize_headers(raw_headers)
            self._validate_required_columns(headers)

            # Read data rows
            rows: list[dict[str, Any]] = []
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                row_dict: dict[str, Any] = {}
                for idx, value in enumerate(row_cells):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = value
                # Skip completely empty rows
                if any(v is not None and str(v).strip() for v in row_dict.values()):
                    rows.append(row_dict)

            wb.close()
            logger.info(
                "inventory.excel_parsed",
                rows_found=len(rows),
                columns=len(headers),
            )
            return rows

        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError(
                f"Failed to parse Excel file: {exc}",
                operation="parse_excel",
            ) from exc

    def _parse_csv(self, file_bytes: bytes) -> list[dict[str, Any]]:
        """Parse CSV file bytes into a list of row dicts.

        Tries UTF-8-BOM, then UTF-8, then latin-1 for encoding.

        Args:
            file_bytes: Raw CSV content.

        Returns:
            List of dicts with normalized column keys.

        Raises:
            ValidationError: If CSV cannot be parsed or
                required columns are missing.
        """
        text: Optional[str] = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            raise ValidationError(
                "Cannot decode CSV file — tried UTF-8 and Latin-1",
                operation="parse_csv",
            )

        try:
            reader = csv.DictReader(io.StringIO(text))
            raw_headers = reader.fieldnames or []
            headers = self._normalize_headers(list(raw_headers))
            self._validate_required_columns(headers)

            # Re-create reader with normalized headers
            reader = csv.DictReader(
                io.StringIO(text), fieldnames=headers
            )
            next(reader)  # Skip the original header row

            rows: list[dict[str, Any]] = []
            for row in reader:
                # Skip completely empty rows
                if any(
                    v is not None and str(v).strip()
                    for v in row.values()
                ):
                    rows.append(dict(row))

            logger.info(
                "inventory.csv_parsed",
                rows_found=len(rows),
                columns=len(headers),
            )
            return rows

        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError(
                f"Failed to parse CSV file: {exc}",
                operation="parse_csv",
            ) from exc

    # ── Column Normalization ────────────────────────────────────────

    def _normalize_headers(
        self, raw_headers: Sequence[str]
    ) -> list[str]:
        """Normalize raw column headers to canonical field names.

        Strips whitespace, lowercases, replaces spaces/hyphens with
        underscores, and applies alias mapping.

        Args:
            raw_headers: Raw header strings from the file.

        Returns:
            List of normalized header names.  Unknown columns get an
            empty string (ignored during row processing).
        """
        normalized: list[str] = []
        for raw in raw_headers:
            clean = raw.strip().lower().replace(" ", "_").replace("-", "_")
            # Apply alias mapping
            if clean in _COLUMN_ALIASES:
                clean = _COLUMN_ALIASES[clean]
            # Only keep known columns
            if clean in KNOWN_COLUMNS:
                normalized.append(clean)
            else:
                normalized.append("")
                if clean:
                    logger.debug(
                        "inventory.unknown_column_ignored",
                        raw_header=raw,
                        normalized=clean,
                    )
        return normalized

    def _validate_required_columns(
        self, headers: Sequence[str]
    ) -> None:
        """Check that all required columns are present.

        Args:
            headers: Normalized header list.

        Raises:
            ValidationError: If any required column is missing.
        """
        present = set(headers)
        missing = REQUIRED_COLUMNS - present
        if missing:
            raise ValidationError(
                f"Missing required columns: {', '.join(sorted(missing))}. "
                f"Required: {', '.join(sorted(REQUIRED_COLUMNS))}",
                operation="validate_columns",
            )

    # ── Row Upsert ──────────────────────────────────────────────────

    async def _upsert_rows(
        self,
        distributor_id: str,
        sync_log_id: str,
        rows: list[dict[str, Any]],
    ) -> SyncResult:
        """Validate and upsert each row into the catalog.

        Processes rows sequentially to preserve ordering and collect
        per-row errors.  Continues past individual row failures to
        maximize data import.

        Args:
            distributor_id: Tenant UUID.
            sync_log_id: Active sync log entry ID.
            rows: Parsed row dicts.

        Returns:
            SyncResult with full counts and errors.
        """
        inserted = 0
        updated = 0
        failed = 0
        errors: list[RowError] = []

        for idx, row in enumerate(rows, start=2):  # Row 2 = first data row
            try:
                create_data = self._validate_row(distributor_id, row, idx)
                item, was_inserted = await catalog_repo.upsert_by_sku_or_name(
                    distributor_id, create_data
                )
                if was_inserted:
                    inserted += 1
                else:
                    updated += 1

            except (ValidationError, DatabaseError) as exc:
                failed += 1
                errors.append(
                    RowError(
                        row_number=idx,
                        error=str(exc),
                    )
                )
                logger.warning(
                    "inventory.row_sync_failed",
                    row_number=idx,
                    error=str(exc),
                )
                continue

        rows_processed = inserted + updated + failed

        # Determine final status
        if failed == 0:
            status = SyncStatus.COMPLETED
        elif inserted + updated > 0:
            status = SyncStatus.PARTIAL
        else:
            status = SyncStatus.FAILED

        # Finalize sync log
        await self._complete_sync_log(
            sync_log_id=sync_log_id,
            status=status,
            rows_processed=rows_processed,
            rows_inserted=inserted,
            rows_updated=updated,
            rows_failed=failed,
            errors=errors,
        )

        # Invalidate catalog cache so subsequent reads see fresh data
        self._catalog_svc.invalidate_cache(distributor_id)

        logger.info(
            "inventory.sync_completed",
            distributor_id=distributor_id,
            status=status.value,
            rows_processed=rows_processed,
            inserted=inserted,
            updated=updated,
            failed=failed,
        )

        return SyncResult(
            sync_log_id=sync_log_id,
            distributor_id=distributor_id,
            status=status,
            rows_processed=rows_processed,
            rows_inserted=inserted,
            rows_updated=updated,
            rows_failed=failed,
            errors=errors,
        )

    def _validate_row(
        self,
        distributor_id: str,
        row: dict[str, Any],
        row_number: int,
    ) -> CatalogItemCreate:
        """Validate a single row and convert to CatalogItemCreate.

        Args:
            distributor_id: Tenant UUID.
            row: Dict of column → value.
            row_number: Row index for error reporting.

        Returns:
            Validated CatalogItemCreate ready for upsert.

        Raises:
            ValidationError: If required fields are missing or have
                invalid types.
        """
        # ── medicine_name (required) ────────────────────────────────
        medicine_name = self._str_value(row.get("medicine_name"))
        if not medicine_name:
            raise ValidationError(
                f"Row {row_number}: medicine_name is required",
                operation="validate_row",
            )

        # ── price_per_unit_paisas (required) ────────────────────────
        price_raw = row.get("price_per_unit_paisas")
        price = self._int_value(price_raw)
        if price is None or price < 0:
            raise ValidationError(
                f"Row {row_number}: price_per_unit_paisas must be a "
                f"non-negative integer, got: {price_raw!r}",
                operation="validate_row",
            )

        # ── stock_quantity (required) ───────────────────────────────
        stock_raw = row.get("stock_quantity")
        stock = self._int_value(stock_raw)
        if stock is None or stock < 0:
            raise ValidationError(
                f"Row {row_number}: stock_quantity must be a "
                f"non-negative integer, got: {stock_raw!r}",
                operation="validate_row",
            )

        # ── Optional fields ─────────────────────────────────────────
        kwargs: dict[str, Any] = {
            "distributor_id": UUID(distributor_id),
            "medicine_name": medicine_name,
            "price_per_unit_paisas": price,
            "stock_quantity": stock,
            "is_in_stock": stock > 0,
        }

        # String fields
        for col in (
            "generic_name", "brand_name", "manufacturer", "category",
            "strength", "unit", "sku", "barcode", "image_url",
        ):
            val = self._str_value(row.get(col))
            if val:
                kwargs[col] = val

        # Form enum
        form_val = self._str_value(row.get("form"))
        if form_val:
            kwargs["form"] = form_val

        # Integer fields
        for col in ("units_per_pack", "mrp_paisas", "low_stock_threshold"):
            val = self._int_value(row.get(col))
            if val is not None and val >= 0:
                kwargs[col] = val

        # Boolean fields
        for col in (
            "allow_order_when_out_of_stock",
            "requires_prescription",
            "is_controlled_substance",
        ):
            val = self._bool_value(row.get(col))
            if val is not None:
                kwargs[col] = val

        # search_keywords — comma-separated string or list
        kw_raw = row.get("search_keywords")
        if kw_raw:
            if isinstance(kw_raw, list):
                kwargs["search_keywords"] = [
                    str(k).strip() for k in kw_raw if str(k).strip()
                ]
            elif isinstance(kw_raw, str) and kw_raw.strip():
                kwargs["search_keywords"] = [
                    k.strip() for k in kw_raw.split(",") if k.strip()
                ]

        try:
            return CatalogItemCreate(**kwargs)
        except Exception as exc:
            raise ValidationError(
                f"Row {row_number}: validation failed — {exc}",
                operation="validate_row",
            ) from exc

    # ── Sync Log Helpers ────────────────────────────────────────────

    async def _start_sync_log(
        self,
        distributor_id: str,
        sync_source: SyncSource,
        file_name: str,
        file_url: Optional[str],
    ) -> InventorySyncLog:
        """Create a sync log entry with status STARTED.

        Args:
            distributor_id: Tenant UUID.
            sync_source: Origin of the sync.
            file_name: Name of the source file.
            file_url: Optional URL to the source file.

        Returns:
            Newly created InventorySyncLog.
        """
        data = InventorySyncLogCreate(
            distributor_id=UUID(distributor_id),
            sync_source=sync_source,
            file_name=file_name,
            file_url=file_url,
            status=SyncStatus.STARTED,
        )
        return await sync_log_repo.create(data)

    async def _complete_sync_log(
        self,
        sync_log_id: str,
        status: SyncStatus,
        rows_processed: int,
        rows_inserted: int,
        rows_updated: int,
        rows_failed: int,
        errors: list[RowError],
    ) -> None:
        """Update sync log entry with final results.

        Args:
            sync_log_id: UUID of the sync log entry.
            status: Final sync status.
            rows_processed: Total rows processed.
            rows_inserted: Rows inserted.
            rows_updated: Rows updated.
            rows_failed: Rows that failed.
            errors: List of per-row errors.
        """
        error_dicts = [
            {
                "row_number": e.row_number,
                "column": e.column,
                "value": e.value,
                "error": e.error,
            }
            for e in errors[:100]  # Cap at 100 errors to avoid huge JSONB
        ]

        update_data = InventorySyncLogUpdate(
            status=status,
            rows_processed=rows_processed,
            rows_inserted=rows_inserted,
            rows_updated=rows_updated,
            rows_failed=rows_failed,
            error_details=error_dicts,
            completed_at=datetime.now(tz=timezone.utc),
        )

        try:
            await sync_log_repo.update(sync_log_id, update_data)
        except Exception as exc:
            # Log but do not raise — sync log failure should not
            # invalidate the actual sync results
            logger.error(
                "inventory.sync_log_update_failed",
                sync_log_id=sync_log_id,
                error=str(exc),
            )

    # ── URL Helpers ─────────────────────────────────────────────────

    @staticmethod
    def _convert_gdrive_url(url: str) -> str:
        """Convert Google Drive share links to direct download URLs.

        Handles these patterns:
            - ``https://drive.google.com/file/d/FILE_ID/...``
            - ``https://drive.google.com/open?id=FILE_ID``

        Args:
            url: Original URL.

        Returns:
            Direct download URL (or the original if not a GDrive link).
        """
        match = _GDRIVE_FILE_ID_RE.search(url)
        if match:
            file_id = match.group(1)
            return (
                f"https://drive.google.com/uc?export=download&id={file_id}"
            )
        return url

    @staticmethod
    def _extract_filename_from_url(url: str) -> str:
        """Best-effort filename extraction from a URL.

        Args:
            url: HTTP(S) URL.

        Returns:
            Filename string or ``"unknown_file"``.
        """
        try:
            path = url.split("?")[0].rstrip("/")
            name = path.rsplit("/", 1)[-1]
            if name and "." in name:
                return name
        except Exception:
            pass
        return "unknown_file"

    # ── Type Coercion Helpers ───────────────────────────────────────

    @staticmethod
    def _str_value(val: Any) -> Optional[str]:
        """Coerce a cell value to a stripped string or None.

        Args:
            val: Raw cell value.

        Returns:
            Stripped string or None if empty/None.
        """
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    @staticmethod
    def _int_value(val: Any) -> Optional[int]:
        """Coerce a cell value to int or None.

        Handles floats (e.g. Excel reads ``100`` as ``100.0``).

        Args:
            val: Raw cell value.

        Returns:
            Integer or None if not convertible.
        """
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _bool_value(val: Any) -> Optional[bool]:
        """Coerce a cell value to bool or None.

        Recognizes: ``True/False``, ``yes/no``, ``1/0``, ``y/n``.

        Args:
            val: Raw cell value.

        Returns:
            Boolean or None if not convertible.
        """
        if val is None:
            return None
        if isinstance(val, bool):
            return val
        s = str(val).strip().lower()
        if s in ("true", "yes", "1", "y"):
            return True
        if s in ("false", "no", "0", "n"):
            return False
        return None


# ── Module-level singleton ──────────────────────────────────────────

inventory_sync_service = InventorySyncService()
